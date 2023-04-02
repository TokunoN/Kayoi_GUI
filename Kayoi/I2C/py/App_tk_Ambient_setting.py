# Ambient の設定を管理するGUIアプリケーション
# tk版

from get_filepath import get_filepath
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox as mb
import pandas as pd

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

class App(ttk.Frame):
    def __init__(self, master = None):
        super().__init__(master)
        self.master.title("App_tk_Ambient_setting")
        # self.master.geometry("100x300")

        self.file = "../resource/AmbientSetting.xlsx"
        self.filepath = get_filepath(self.file)
        try:
            self.wb = openpyxl.load_workbook(self.filepath)
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()
            
            chInfo_sheet = self.wb.create_sheet("chInfo")
            df = pd.DataFrame(columns=["chNumber", "chID", "writeKey", "readKey", "userKey"])
            df["chNumber"] = [f"ch{i}" for i in range(1,9)]
            for r in dataframe_to_rows(df, index=False, header=True):
                chInfo_sheet.append(r)
            
            df = pd.DataFrame(columns=["dataID", "house_name", "raspi_name", "sensor_name", "datatype"])
            df["dataID"] = [f"d{i}" for i in range(1,9)]
            for i in range(1,9):
                sheet = self.wb.create_sheet(f"ch{i}")
                for r in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(r)
            
            self.wb.remove(self.wb["Sheet"])
            self.wb.save(self.filepath)

        self.sheetname_list = self.wb.sheetnames

        for sheetname in self.sheetname_list:
            ttk.Button(self.master, text=f"{sheetname}", command=lambda sheetname = sheetname: self.open_form(sheetname)).pack(padx = 5, pady = 1)


    def open_form(self, sheetname):
        self.wb = openpyxl.load_workbook(self.filepath)
        self.form_window = FormWindow(self, self.wb, sheetname)
        self.form_window.grab_set()
        self.form_window.focus_set()
        self.form_window.protocol("WM_DELETE_WINDOW", self.form_window.ask_save)


class FormWindow(tk.Toplevel):
    def __init__(self, master, wb, sheetname):
        super().__init__(master=None)
        self.title("Form_Window")
        # self.geometry("700x400")

        self.wb = wb
        self.ws = self.wb[f"{sheetname}"]
        g_all = self.ws.values
        self.sv_list = [[tk.StringVar(
            value=f"{cell}" if cell != None else "") for cell in row] for row in g_all]
        self.entry_list = [[tk.Entry(
            self,
            textvariable=sv,
            font = "NotoSansJP",
            state="readonly" if i*j == 0 else "normal",
            justify="center" if i*j == 0 else "left"
        ) for j, sv in enumerate(row)
        ] for i, row in enumerate(self.sv_list)]
        [[entry.grid(
            row=i,
            column=j,
            padx = [max(0, (1 - j)) * 5, max(0, (len(row) * (-1) + j + 2)) * 5],
            pady = [max(0, (1 - i)) * 5, max(0, (len(self.entry_list) * (-1) + i + 2)) * 5]
        ) for j, entry in enumerate(row)]
         for i, row in enumerate(self.entry_list)]
        
    def save(self):
        # 入力されたテキストをファイルに保存する関数
        l_2d = [[sv.get() for sv in row] for row in self.sv_list]
        for i, row in enumerate(l_2d):
            for j, cell in enumerate(row):
                self.ws.cell(row=i+1, column = j+1, value = cell)
        
        self.file = "../resource/AmbientSetting.xlsx"
        self.filepath = get_filepath(self.file)
        self.wb.save(self.filepath)
        self.destroy()

    def ask_save(self):
        # ウィンドウを閉じる前に保存するかどうかを尋ねる関数
        answer = mb.askyesnocancel("Save", "保存して終了しますか？")
        if answer:
            self.save()
        elif answer == False:
            self.destroy()
        else:
            pass

root = tk.Tk()
app = App(master=root)
app.mainloop()