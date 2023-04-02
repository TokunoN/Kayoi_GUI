# Ambientの設定を行うGUIアプリケーション

import flet as ft
import pandas as pd
from Read_csv import Read_csv
from get_filepath import get_filepath
import math
import os

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import re

class Ambient_Table(ft.UserControl):
    def __init__(self, sheetname):
        super().__init__()
        self.sheetname = sheetname
        self.sheet = wb[self.sheetname]
        self.path = os.path.dirname(__file__)
        self.df = self.convert_excelsheet_into_2dlist()
        self.file = "AmbientSetting.xlsx"

    def convert_excelsheet_into_2dlist(self):
        g_all = self.sheet.values

        l_2d = [[cell for cell in row] for row in g_all]
        df = pd.DataFrame(l_2d, columns = l_2d[0])
        df.drop([0], axis = 0, inplace=True)
        return df

    def select_changed(self, e):
        e.control.selected = not e.control.selected
        self.update()
    
    def get_row_and_column_num_in_selected_cell(self, target):
        # eのtargetについて(2023/2/2現在)
        # DataTableのセルをクリックしたときのevent(=e)のtargetを参照して
        # Datatableのセルの位置を計算している。
        # おそらくtargetはcontrolの数に依って変動するので、あまりよくない方法
        # ほかに代替案があれば変更するべき。ただし代替案が分からない。

        # controlの追加・削除等を行った際は設定している数値を再度変更すること
        # ch1 と ch2 の最初の行のhouse_name列のセルをクリックした時のtargetの値から2を引いたものをchx_first_cell_numに代入すると動く。
        # 
        # chInfo:
        # 20,22,24,26,28
        # 31,33,35,37,39
        # ...
        # 97,99,101,103,105
        # 
        # ch1:
        # 123,125,127,129,131
        # 134,136,138,140,142
        # ...
        # 200,202,204,206,208
        # 
        # ch2:
        # 226,228,230,232,234
        # 237,239
        # 
        # 
        ch1_first_cell_num = 123
        ch2_first_cell_num = 226
        btw_tab_num = ch2_first_cell_num - ch1_first_cell_num

        if target < ch1_first_cell_num:
            ch_num = 0
            tab_offset = 19
            row_increment = 11
            column_increment = 2
        else:
            ch_num = (target - ch1_first_cell_num) // btw_tab_num + 1
            tab_offset = ch1_first_cell_num + btw_tab_num * (ch_num - 1)
            row_increment = 11
            column_increment = 2

        target_norm = target - tab_offset
        row_num = target_norm // row_increment
        column_num = int(target_norm % row_increment / column_increment)

        return row_num, column_num

    def edit_tapped(self, e):
        # タップしたときに編集画面を提示
        target = int(e.target[1:])
        self.row_num, self.column_num = self.get_row_and_column_num_in_selected_cell(target)
        dc = self.dt.rows[self.row_num].cells[self.column_num]
        dc.show_edit_icon = False
        dc.on_tap = None

        tf = ft.TextField(
            value=dc.content.value,
            label = self.dt.columns[self.column_num].label.value,
            autofocus=True,
            on_submit=self.submit_entered,
            on_blur=self.submit_entered,
            width=300
            )
        dc.content = tf        
        self.update()

    def submit_entered(self, e):
        # フォーカスが外れた=編集が終了した -> 通常の表示に戻す 
        # この関数上ではeのtargetは、選択するごとに加算される仕組みで、
        # セルの座標を取得することは不可能
        # そのため、静的変数として保存したrow_numとcolumn_numを使用する。
        dc = self.dt.rows[self.row_num].cells[self.column_num]
        dc.content = ft.Text(value=dc.content.value)
        dc.show_edit_icon = True
        dc.on_tap = self.edit_tapped
        
        self.update()

    def save_dt_to_xlsx(self, e):
        # データテーブルをエクセルファイルに保存する
        # 問題提起
        # 数値が入力されても文字列として保存される
        # 数値を使用する場面はあるか？
        # チャネルIDは使用時にstrに変換されるので問題なさそう。
        df_for_xlsx_sheet = pd.DataFrame(
            data = [[cell.content.value  for cell in row.cells] for row in self.dt.rows],
            columns = [col.label.value for col in self.dt.columns]
        )

        if os.path.isfile(filepath):
            with pd.ExcelWriter(filepath, mode = "a", if_sheet_exists='replace') as writer:
                df_for_xlsx_sheet.to_excel(writer, sheet_name=self.sheetname, index=False)
        else:
            df_for_xlsx_sheet.to_excel(filepath,sheet_name=self.sheetname, index=False)

    def convert_df_cell_into_datacell(self, x):
        # pd.DataFrameをflet.DatatableのDatacellに変換する
        # セルの中身がNoneかnanなら空白を返す
        value = ""
        if isinstance(x, str):
            value = x
        elif x == None:
            pass
        elif not math.isnan(x):
            value = x
        else:
            pass
        return ft.DataCell(ft.Text(value, expand=True), show_edit_icon=True, on_tap=self.edit_tapped)

    def build(self):
        # GUI本体
        self.dt = ft.DataTable(show_checkbox_column=True)
        self.dt.columns = (list(map(lambda x: ft.DataColumn(ft.Text(
            x, width=300 if self.sheetname == "chInfo" else 200)), self.df.columns)))
        self.dt.columns[0].label.width = None
        for index in self.df.index:
            dc = list(map(self.convert_df_cell_into_datacell,
                      self.df.loc[index, :]))
            dc[0].show_edit_icon = False
            dc[0].on_tap = None
            dr = ft.DataRow(cells=dc, selected=False)
            dr.on_select_changed = self.select_changed
            self.dt.rows.append(dr)

        return ft.Column([
            ft.Row([self.dt], scroll="adaptive"),
            # save_btn
        ], horizontal_alignment="center", scroll="adaptive")


def main(page: ft.Page):
    # 日本語フォントファイルの指定
    font_file = "assets/fonts/NotoSansJP-Light.otf"
    font_filepath = get_filepath(font_file)
    if os.path.exists(font_filepath) == False:
        message = "日本語表示用のフォントファイルが存在しません"
        file = "../stderr.log"
        err_filepath = get_filepath(file)
        with open(err_filepath, mode = "a") as f:
            f.write(message + "\n")
        return
    page.fonts = {
            "NotoSansJP": font_filepath
    }

    # page.theme はライトモードのテーマ
    # page.dark_theme はダークモードのテーマ
    # https://github.com/flet-dev/flet/issues/544
    # どちらが選択されるか分からないので両方記述しておく。
    page.theme = ft.Theme(font_family="NotoSansJP")
    page.dark_theme = ft.Theme(font_family="NotoSansJP")

    def window_event(e):
        # ウィンドウを閉じたときにダイアログを表示するように設定
        if e.data == "close":
            page.dialog = close_confirm_dialog
            close_confirm_dialog.open = True
            page.update()

    def save_yes_click(e):
        # 保存する場合の挙動
        for i, tab in enumerate(t.tabs):
            if i == 0:
                tab.content.controls[0].save_dt_to_xlsx(e)
            else:
                tab.content.save_dt_to_xlsx(e)

        page.window_destroy()

    def save_no_click(e):
        # 保存しないときはそのまま閉じる
        page.window_destroy()

    def close_no_click(e):
        # キャンセル: ダイアログを閉じる
        close_confirm_dialog.open = False
        page.update()
    
    # ウィンドウを閉じるときにイベントが発生するように設定
    page.window_prevent_close = True
    page.on_window_event = window_event

    close_confirm_dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text("アプリを終了します"),
        content=ft.Text("保存していない変更を保存しますか？"),
        actions=[
            ft.ElevatedButton("終了しない", on_click=close_no_click),
            ft.ElevatedButton("変更を破棄して終了", on_click=save_no_click),
            ft.OutlinedButton("変更を保存して終了", on_click=save_yes_click),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    # windowのサイズをグローバル宣言
    # Table()の中からwindowのサイズを取得できる方法が見つかれば必要ない
    global window_width
    window_width = page.window_width

    # エクセルファイルの読み込み
    # 無ければ作成する
    global wb
    file = "../resource/AmbientSetting.xlsx"
    global filepath
    filepath = get_filepath(file)
    try:
        wb = openpyxl.load_workbook(filepath)
    except FileNotFoundError:
        
        wb = openpyxl.Workbook()
        
        chInfo_sheet = wb.create_sheet("chInfo")
        df = pd.DataFrame(columns=["chNumber", "chID", "writeKey", "readKey", "userKey"])
        df["chNumber"] = [f"ch{i}" for i in range(1,9)]
        for r in dataframe_to_rows(df, index=False, header=True):
            chInfo_sheet.append(r)
        
        df = pd.DataFrame(columns=["dataID", "house_name", "raspi_name", "sensor_name", "datatype"])
        df["dataID"] = [f"d{i}" for i in range(1,9)]
        for i in range(1,9):
            sheet = wb.create_sheet(f"ch{i}")
            for r in dataframe_to_rows(df, index=False, header=True):
                sheet.append(r)
        
        wb.remove(wb["Sheet"])

    pattern = re.compile(r"ch\d")


    # タブで表示するための準備
    tabs_list = [
        ft.Tab(
            text="chInfo",
            content=ft.Column([
                Ambient_Table(sheetname="chInfo"),
            ]),
            icon=ft.icons.SETTINGS
        )
    ]

    chs_list = [sheet for sheet in wb.sheetnames if pattern.match(sheet)]
    icon_list = [ft.icons.FILTER_1, ft.icons.FILTER_2, ft.icons.FILTER_3, ft.icons.FILTER_4,
                 ft.icons.FILTER_5, ft.icons.FILTER_6, ft.icons.FILTER_7, ft.icons.FILTER_8]

    tabs_list = tabs_list + [
        ft.Tab(
            text=ch,
            content=Ambient_Table(sheetname=ch)
        )
        for ch in chs_list]

    # 数字アイコンの設定
    for i in range(1,9):
        tabs_list[i].icon = icon_list[i-1]

    t = ft.Tabs(
        selected_index=0,
        animation_duration=300,
        tabs=tabs_list,
        expand=1,
    )
    
    page.add(t)

ft.app(target=main, assets_dir="assets")